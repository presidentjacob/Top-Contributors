from bs4 import BeautifulSoup
import requests
import openpyxl

# Set all urls
year_2022 = 'https://openaccess.thecvf.com/CVPR2022?day=all'
year_2023 = 'https://openaccess.thecvf.com/CVPR2023?day=all'
year_2024 = 'https://openaccess.thecvf.com/CVPR2024?day=all'

def get_url(url):
    response = requests.get(url)
    # Check response if the response is successful
    if response.status_code == 200:
        return get_all(response)
    else:
        # If response is not 200, raise an exception with the status code and return None
        raise Exception(f'Response status code: {response.status_code}')
    return None

def get_all(response):
    # Create soup to find all people
    soup = BeautifulSoup(response.text, 'lxml')
    find_people = soup.find_all('form', class_='authsearch')
    # Create a dictionary to store all people
    all_people = dict()

    # Loop through all people and add them to the dictionary
    for people in find_people:
        for person in people.find_all('a'):
            name = person.text.strip()

            # Add to dictionary if name doesn't exist
            if name not in all_people:
                all_people[name] = 1
            # If name exists, increment the count
            else:
                all_people[name] += 1

    return all_people

def get_top_three(all_2022, all_2023, all_2024):
    all_total = dict()
    # Combine all years into one dictionary
    for person, total in all_2022.items():
        if person not in all_total:
            all_total[person] = total
        else:
            all_total[person] += total

    for person, total in all_2023.items():
        if person not in all_total:
            all_total[person] = total
        else:
            all_total[person] += total
    
    for person, total in all_2024.items():
        if person not in all_total:
            all_total[person] = total
        else:
            all_total[person] += total

    '''
    Sort the dictionary by value and get the top three
    Use lambda function to sort by value in descending order
    x[1] gets the value to sort by, the total contributions of each person
    Return the top three people with the most contributions
    Reverse as it sorts in ascending order by default
    Use [:3] to get the top three people
    '''

    top_three = sorted(all_total.items(), key=lambda x: x[1], reverse=True)[:3]
    return top_three

def add_to_file(top_three):
    # Create a workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Top Three Contributors'

    # Add headers to the sheet
    for i in range(2, 5):
        ws.cell(row=i, column=1, value=f'202{i}')
    ws.cell(row=5, column=1, value='Total')
    
    wb.save('top_three_contributors.xlsx')
    

def main():
    # Get all people from each year
    all_2022 = get_url(year_2022)
    all_2023 = get_url(year_2023)
    all_2024 = get_url(year_2024)

    # Get the top three people with the most contributions
    top_three = get_top_three(all_2022, all_2023, all_2024)
    add_to_file(top_three)

main()